"""
洗衣機顧客名單清洗工具

功能：
  1. 把跨欄置中的日期欄拆開，每筆訂單都補上對應日期
  2. 把含「休假」的列直接刪除（針對 2016~2021 舊格式）

支援格式：
  .xls  (2016~2021)：日期欄在 B 欄(index 1)，有休假列要刪
  .xlsx (2022~)：日期欄在 A 欄(index 0)，無休假列

輸入：拖檔（多檔）或命令列傳入路徑
輸出：同目錄底下 cleaned/<原檔名>_cleaned.xlsx
"""
from __future__ import annotations
import sys
from pathlib import Path
from datetime import datetime, time

import xlrd
from xlrd import xldate
from openpyxl import Workbook, load_workbook


SKIP_TOKEN = "休假"


def xlrd_value(sh, r, c, datemode):
    """讀取 xlrd cell，把 Excel date 序號轉成 datetime。"""
    ct = sh.cell_type(r, c)
    v = sh.cell_value(r, c)
    if ct == xlrd.XL_CELL_DATE:
        try:
            t = xldate.xldate_as_tuple(v, datemode)
            if t[:3] == (0, 0, 0):
                return time(*t[3:])
            return datetime(*t)
        except Exception:
            return v
    if ct == xlrd.XL_CELL_BOOLEAN:
        return bool(v)
    if ct == xlrd.XL_CELL_EMPTY:
        return None
    if ct == xlrd.XL_CELL_NUMBER:
        # 整數就用 int，避免一堆 .0
        if isinstance(v, float) and v.is_integer():
            return int(v)
    return v


def row_should_skip(values: list, merge_row_values: set) -> bool:
    """判斷整列是否該刪除（含休假字樣）。"""
    for v in values:
        if isinstance(v, str) and SKIP_TOKEN in v:
            return True
    for v in merge_row_values:
        if isinstance(v, str) and SKIP_TOKEN in v:
            return True
    return False


def clean_xls(input_path: Path, output_path: Path) -> dict:
    """處理 .xls：日期欄在 col 1(B)，刪除休假列。"""
    DATE_COL = 1  # 0-based, B 欄
    wb = xlrd.open_workbook(str(input_path), formatting_info=True)
    new_wb = Workbook()
    new_wb.remove(new_wb.active)

    stats = {"sheets": 0, "rows_in": 0, "rows_out": 0,
             "skipped": 0, "filled": 0, "no_date": []}

    for sh in wb.sheets():
        stats["sheets"] += 1
        # 建立 (r,c) -> 拆開後該填的值
        merge_value_for: dict[tuple[int, int], object] = {}
        # 每列被合併涵蓋的 top-left 值（用來判斷休假整列合併）
        row_merge_values: dict[int, list] = {}
        for rlo, rhi, clo, chi in sh.merged_cells:
            top_val = xlrd_value(sh, rlo, clo, wb.datemode)
            for r in range(rlo, rhi):
                row_merge_values.setdefault(r, []).append(top_val)
                for c in range(clo, chi):
                    if (r, c) != (rlo, clo):
                        merge_value_for[(r, c)] = top_val

        new_sh = new_wb.create_sheet(sh.name[:31] or f"Sheet{stats['sheets']}")
        out_r = 1
        for r in range(sh.nrows):
            stats["rows_in"] += 1
            row_vals = []
            for c in range(sh.ncols):
                if (r, c) in merge_value_for:
                    row_vals.append(merge_value_for[(r, c)])
                else:
                    row_vals.append(xlrd_value(sh, r, c, wb.datemode))

            if row_should_skip(row_vals, set(filter(None, row_merge_values.get(r, [])))):
                stats["skipped"] += 1
                continue

            # 檢查非空白資料列日期是否缺漏
            has_data = any(v not in (None, "", 0) for i, v in enumerate(row_vals) if i != DATE_COL)
            if has_data and row_vals[DATE_COL] in (None, ""):
                stats["no_date"].append(f"{sh.name}!row{r+1}")

            for c, v in enumerate(row_vals, start=1):
                if (r, c - 1) in merge_value_for:
                    stats["filled"] += 1
                new_sh.cell(row=out_r, column=c, value=v)
            out_r += 1
            stats["rows_out"] += 1

    new_wb.save(str(output_path))
    return stats


def clean_xlsx(input_path: Path, output_path: Path) -> dict:
    """處理 .xlsx：日期欄在 col 0(A)，沒有休假列。"""
    DATE_COL = 1  # 1-based for openpyxl
    wb = load_workbook(str(input_path))
    stats = {"sheets": 0, "rows_in": 0, "rows_out": 0,
             "skipped": 0, "filled": 0, "no_date": []}

    for sh in wb.worksheets:
        stats["sheets"] += 1
        stats["rows_in"] += sh.max_row

        # snapshot ranges 後再 unmerge（直接 iterate 同時改會炸）
        ranges = list(sh.merged_cells.ranges)
        for mr in ranges:
            top_val = sh.cell(mr.min_row, mr.min_col).value
            sh.unmerge_cells(str(mr))
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    if (r, c) != (mr.min_row, mr.min_col):
                        sh.cell(row=r, column=c).value = top_val
                        stats["filled"] += 1

        # 檢查日期缺漏（從 row 2 開始，假設 row 1 是 header）
        start_row = 2 if any(isinstance(sh.cell(1, c).value, str)
                              for c in range(1, sh.max_column + 1)) else 1
        for r in range(start_row, sh.max_row + 1):
            row_vals = [sh.cell(r, c).value for c in range(1, sh.max_column + 1)]
            has_data = any(v not in (None, "", 0)
                           for i, v in enumerate(row_vals, start=1) if i != DATE_COL)
            if has_data and sh.cell(r, DATE_COL).value in (None, ""):
                stats["no_date"].append(f"{sh.title}!row{r}")

        stats["rows_out"] += sh.max_row

    wb.save(str(output_path))
    return stats


def process_one(path: Path) -> tuple[Path, dict] | tuple[None, str]:
    suffix = path.suffix.lower()
    if suffix not in (".xls", ".xlsx"):
        return None, f"略過（非 Excel 檔）：{path.name}"
    if not path.exists():
        return None, f"找不到檔案：{path}"

    out_dir = path.parent / "cleaned"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f"{path.stem}_cleaned.xlsx"

    if suffix == ".xls":
        stats = clean_xls(path, out_path)
    else:
        stats = clean_xlsx(path, out_path)
    return out_path, stats


def main():
    args = sys.argv[1:]
    if not args:
        print("用法：python clean_orders.py <檔1> [檔2] ...")
        print("或：把檔案拖到 clean.bat 上")
        input("按 Enter 結束...")
        return

    print(f"=== 清洗 {len(args)} 個檔案 ===\n")
    ok, fail = 0, 0
    for arg in args:
        p = Path(arg)
        print(f"處理：{p.name}")
        try:
            out, info = process_one(p)
            if out is None:
                print(f"  ! {info}")
                fail += 1
                continue
            print(f"  -> {out}")
            print(f"     sheets={info['sheets']}  "
                  f"輸入列={info['rows_in']}  "
                  f"輸出列={info['rows_out']}  "
                  f"刪除休假列={info['skipped']}  "
                  f"補日期={info['filled']}")
            if info["no_date"]:
                print(f"     ⚠ 原始檔有 {len(info['no_date'])} 列缺日期（不在合併範圍內，需手動補）:")
                for loc in info["no_date"][:10]:
                    print(f"        - {loc}")
                if len(info["no_date"]) > 10:
                    print(f"        ... 還有 {len(info['no_date']) - 10} 筆")
            ok += 1
        except Exception as e:
            print(f"  ! 失敗：{type(e).__name__}: {e}")
            fail += 1
        print()

    print(f"=== 完成：成功 {ok}，失敗 {fail} ===")
    if sys.stdin.isatty():
        input("按 Enter 結束...")


if __name__ == "__main__":
    main()
