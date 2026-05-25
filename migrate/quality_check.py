"""
資料品質檢查：匯入前最後一道把關。

檢查項目：
  A. 同電話分成多筆客戶（含副電話）— 應該整合卻沒整合
  B. 同地址分成多筆客戶 — 可能是同一人或夫妻
  C. 姓名欄位異常（含電話號碼、特殊符號、長度異常）
  D. 電話欄位異常（含中文字、長度異常）
  E. 地址欄位異常（純數字、太短、含電話）
  F. 訂單金額異常（負值、超大、為零）

輸出 quality_report.xlsx，每個檢查項一張 sheet。
"""
from __future__ import annotations
import csv
import re
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "out"
REPORT = OUT / "quality_report.xlsx"


def load_csv(name):
    with open(OUT / name, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def main():
    print("讀 csv...")
    customers = load_csv("customers.csv")
    phones = load_csv("customer_phones.csv")
    addresses = load_csv("customer_addresses.csv")
    orders = load_csv("orders.csv")

    cust_by_id = {c["id"]: c for c in customers}
    print(f"  {len(customers)} 客戶 / {len(phones)} 電話 / {len(addresses)} 地址 / {len(orders)} 訂單\n")

    issues = {}

    # ─── A. 同電話分成多筆客戶 ─────────────────────────────────────────────
    print("A. 檢查同電話多客戶...")
    phone_to_custs = defaultdict(list)
    for p in phones:
        digits = re.sub(r"\D", "", p["phone"])
        if len(digits) >= 7:
            phone_to_custs[digits].append({
                "customer_id": p["customer_id"],
                "phone_raw": p["phone"],
                "is_primary": p["is_primary"],
            })
    dup_phones = []
    for phone, items in phone_to_custs.items():
        unique_custs = {x["customer_id"] for x in items}
        if len(unique_custs) > 1:
            for it in items:
                c = cust_by_id.get(it["customer_id"], {})
                dup_phones.append({
                    "normalized_phone": phone,
                    "customer_code": c.get("code", "?"),
                    "customer_name": c.get("name", "?"),
                    "phone_raw": it["phone_raw"],
                    "is_primary": it["is_primary"],
                    "customer_id": it["customer_id"][:8],
                })
    print(f"  找到 {len(dup_phones)} 列（涉及 {len({d['normalized_phone'] for d in dup_phones})} 個共用電話）")
    issues["A_同電話多客戶"] = dup_phones

    # ─── B. 同地址分成多筆客戶 ─────────────────────────────────────────────
    print("B. 檢查同地址多客戶...")
    addr_to_custs = defaultdict(list)
    for a in addresses:
        if a["county"] == "未分類":
            continue
        key = (a["county"], a["district"], a["address"].strip())
        addr_to_custs[key].append({
            "customer_id": a["customer_id"],
            "address_label": a["label"],
        })
    dup_addrs = []
    for key, items in addr_to_custs.items():
        unique_custs = {x["customer_id"] for x in items}
        if len(unique_custs) > 1:
            for it in items:
                c = cust_by_id.get(it["customer_id"], {})
                dup_addrs.append({
                    "county": key[0],
                    "district": key[1],
                    "address": key[2],
                    "customer_code": c.get("code", "?"),
                    "customer_name": c.get("name", "?"),
                    "customer_phone": c.get("phone", "?"),
                    "customer_id": it["customer_id"][:8],
                })
    print(f"  找到 {len(dup_addrs)} 列（涉及 {len({(d['county'], d['district'], d['address']) for d in dup_addrs})} 個共用地址）")
    issues["B_同地址多客戶"] = dup_addrs

    # ─── C. 姓名欄位異常 ─────────────────────────────────────────────────
    print("C. 檢查姓名異常...")
    bad_names = []
    for c in customers:
        name = c["name"]
        reasons = []
        if not name or name == "（舊資料-無姓名）":
            continue  # 已知缺失，跳過
        # 過長
        if len(name) > 15:
            reasons.append("過長 (>15 字)")
        # 含太多數字
        digits = sum(1 for ch in name if ch.isdigit())
        if digits >= 5:
            reasons.append(f"含 {digits} 個數字（疑似電話）")
        # 含 @ 或 http
        if "@" in name or "http" in name.lower():
            reasons.append("含 @ 或 http")
        # 全是符號或空白
        if not any(ch.isalnum() or '一' <= ch <= '鿿' for ch in name):
            reasons.append("無有效字元")
        if reasons:
            bad_names.append({
                "customer_code": c["code"],
                "name": name,
                "phone": c["phone"],
                "原因": "；".join(reasons),
            })
    print(f"  找到 {len(bad_names)} 筆姓名異常")
    issues["C_姓名異常"] = bad_names

    # ─── D. 電話欄位異常 ─────────────────────────────────────────────────
    print("D. 檢查電話異常...")
    bad_phones = []
    for p in phones:
        phone = p["phone"]
        digits = re.sub(r"\D", "", phone)
        reasons = []
        if len(digits) < 7:
            reasons.append(f"太短 ({len(digits)} 碼)")
        if len(digits) > 11:
            reasons.append(f"太長 ({len(digits)} 碼)")
        # 含中文（不該）
        if re.search(r"[一-鿿]", phone):
            reasons.append("含中文字")
        # 全是同一個數字（如 1234567890）
        if len(set(digits)) <= 2 and len(digits) >= 7:
            reasons.append("數字單調（疑似測試）")
        if reasons:
            c = cust_by_id.get(p["customer_id"], {})
            bad_phones.append({
                "customer_code": c.get("code", "?"),
                "customer_name": c.get("name", "?"),
                "phone": phone,
                "is_primary": p["is_primary"],
                "原因": "；".join(reasons),
            })
    print(f"  找到 {len(bad_phones)} 筆電話異常")
    issues["D_電話異常"] = bad_phones

    # ─── E. 地址欄位異常 ─────────────────────────────────────────────────
    print("E. 檢查地址異常...")
    bad_addrs = []
    for a in addresses:
        addr = a["address"]
        reasons = []
        if a["county"] == "未分類" or a["district"] == "未分類":
            reasons.append(f"縣市/鄉鎮未解析（{a['county']}/{a['district']}）")
        if len(addr) < 5:
            reasons.append(f"地址太短 ({len(addr)} 字)")
        # 純數字（含 - 號）
        if re.fullmatch(r"[\d\-\s]+", addr):
            reasons.append("純數字（疑似電話誤入）")
        # 含可能電話 pattern
        if re.search(r"09\d{8}|0\d-?\d{6,8}", addr):
            reasons.append("含電話號碼 pattern")
        # 沒「號」「樓」「巷」「弄」之類的房屋指示
        if not re.search(r"號|樓|室|巷|弄|段|村|里", addr):
            reasons.append("無『號/樓/巷/弄/段/村/里』指示")
        if reasons:
            c = cust_by_id.get(a["customer_id"], {})
            bad_addrs.append({
                "customer_code": c.get("code", "?"),
                "customer_name": c.get("name", "?"),
                "county": a["county"],
                "district": a["district"],
                "address": addr,
                "原因": "；".join(reasons),
            })
    print(f"  找到 {len(bad_addrs)} 筆地址異常")
    issues["E_地址異常"] = bad_addrs

    # ─── F. 訂單金額異常 ─────────────────────────────────────────────────
    print("F. 檢查訂單金額異常...")
    bad_orders = []
    for o in orders:
        try:
            sub = float(o["subtotal"])
            tot = float(o["total"])
        except (ValueError, KeyError):
            continue
        reasons = []
        if tot == 0 and sub == 0:
            reasons.append("金額為 0（原始檔可能漏填）")
        if tot < 0:
            reasons.append(f"金額為負 ({tot})")
        if tot > 100000:
            reasons.append(f"金額過大 (${tot:,.0f})")
        if reasons:
            bad_orders.append({
                "order_code": o["order_code"],
                "subtotal": f"{sub:,.0f}",
                "total": f"{tot:,.0f}",
                "原因": "；".join(reasons),
                "note": o["note"][:60],
            })
    print(f"  找到 {len(bad_orders)} 筆訂單異常")
    issues["F_訂單金額異常"] = bad_orders

    # ─── 輸出 quality_report.xlsx ─────────────────────────────────────────
    print(f"\n寫出 {REPORT.name}...")
    wb = Workbook()
    wb.remove(wb.active)

    # 摘要 sheet
    sh = wb.create_sheet("摘要")
    sh.append(["檢查項", "問題數", "說明"])
    descriptions = {
        "A_同電話多客戶": "同一支電話被分到多個客戶 → 應該整合成 1 個",
        "B_同地址多客戶": "同一個地址被分到多個客戶 → 可能是同人或夫妻",
        "C_姓名異常": "姓名過長/含數字/特殊符號 → 可能是誤填或姓名空白",
        "D_電話異常": "電話過短/過長/含中文 → 可能是欄位錯位",
        "E_地址異常": "地址沒『號/樓/巷』/含電話 → 可能是欄位錯位",
        "F_訂單金額異常": "金額為 0/負/過大 → 原始檔資料品質問題",
    }
    for k, items in issues.items():
        sh.append([k, len(items), descriptions.get(k, "")])
    style_sheet(sh, 3)
    sh.column_dimensions["A"].width = 24
    sh.column_dimensions["C"].width = 50

    # 各檢查 sheet
    for name, items in issues.items():
        if not items:
            continue
        sh = wb.create_sheet(name)
        cols = list(items[0].keys())
        sh.append(cols)
        for it in items:
            sh.append([it.get(c, "") for c in cols])
        style_sheet(sh, len(cols))

    wb.save(str(REPORT))
    print(f"  完成\n")
    print("=" * 60)
    print("摘要：")
    for k, items in issues.items():
        print(f"  {k}: {len(items)}")


def style_sheet(sh, n_cols):
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="EEEEEE")
    for c in range(1, n_cols + 1):
        cell = sh.cell(1, c)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    sh.freeze_panes = "A2"
    for c in range(1, n_cols + 1):
        sh.column_dimensions[get_column_letter(c)].width = 18


if __name__ == "__main__":
    main()
