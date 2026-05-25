"""
Step 1: 讀 8 個 cleaned xlsx，把每列正規化成統一格式輸出 rows.csv。

每個 row 一筆訂單明細（多廠牌的還沒拆，由 build.py 處理）。
"""
from __future__ import annotations
import re
import csv
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook


CLEANED_DIR = Path(r"C:\RenStudio\case\washinmachine\顧客名單\cleaned")
OUT_DIR = Path(__file__).parent / "out"
OUT_DIR.mkdir(exist_ok=True)
OUT_CSV = OUT_DIR / "rows.csv"


# ---------------------------------------------------------------------------
# 格式偵測：3 種版本的 column layout
# ---------------------------------------------------------------------------
# A (2016~2021 .xls cleaned):
#   col0=編號, col1=日期, col2=姓名, col3=電話, col4=地址,
#   col5=廠牌, col6=金額, col7=來源處
# B (2022, 2023/2月+):
#   col0=日期, col1=姓名, col2=電話, col3=地址, col4=項目,
#   col5=廠牌, col6=付款方式, col7=金額, [col13=機器編號]
# C (2023/1月，無 header):
#   col0=日期, col1=姓名, col2=電話, col3=地址,
#   col4=廠牌, col5=金額, col6=來源, col7=機器編號
COL_MAPS = {
    "A": dict(date=1, name=2, phone=3, addr=4, brand=5, amount=6, source=7,
              item=None, payment=None, machine_id=None),
    "B": dict(date=0, name=1, phone=2, addr=3, item=4, brand=5, payment=6,
              amount=7, source=None, machine_id=13),
    "C": dict(date=0, name=1, phone=2, addr=3, brand=4, amount=5, source=6,
              item=None, payment=None, machine_id=7),
}


def detect_format(sheet) -> tuple[str, int]:
    r1 = [sheet.cell(1, c).value for c in range(1, min(11, sheet.max_column + 1))]
    has_header = any(isinstance(v, str) and v in ("日期", "姓名", "編號") for v in r1)
    has_item = any(isinstance(v, str) and "項目" in v for v in r1)

    data_start = 2 if has_header else 1
    if sheet.max_row < data_start:
        return "A", data_start

    rd = [sheet.cell(data_start, c).value for c in range(1, min(11, sheet.max_column + 1))]
    # 找第一個有效資料列
    while data_start <= sheet.max_row and all(v in (None, "") for v in rd):
        data_start += 1
        if data_start > sheet.max_row:
            break
        rd = [sheet.cell(data_start, c).value for c in range(1, min(11, sheet.max_column + 1))]

    if isinstance(rd[0], int) and 0 < rd[0] < 999:
        return "A", data_start
    if has_item:
        return "B", data_start
    return "C", data_start


# ---------------------------------------------------------------------------
# 欄位正規化
# ---------------------------------------------------------------------------
def to_date_str(v, fallback_year: int | None = None) -> str:
    """轉成 YYYY-MM-DD；不行就回空字串。"""
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        # Excel serial number
        try:
            dt = datetime(1899, 12, 30) + __import__("datetime").timedelta(days=int(v))
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return ""
    if isinstance(v, str):
        s = v.strip()
        # "7月24日"
        m = re.match(r"^(\d{1,2})月(\d{1,2})日?$", s)
        if m and fallback_year:
            try:
                return datetime(fallback_year, int(m.group(1)), int(m.group(2))).strftime("%Y-%m-%d")
            except ValueError:
                return ""
        # "2019/8/24" or "2019-08-24"
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return ""


PHONE_SEP = re.compile(r"[\s\n、,;/]+")

def _split_concatenated_phones(digits: str) -> list[str]:
    """把沒分隔符的長字串嘗試拆成多支電話（>=12 碼才拆）。
    例：'09728340900984370217' → ['0972834090', '0984370217']
        '0287923633258' → ['0287923633', '258']（殘留太短會丟掉）
    """
    if len(digits) <= 11:
        return [digits]
    result = []
    rest = digits
    while len(rest) >= 9:
        if rest.startswith("09") and len(rest) >= 10:
            result.append(rest[:10])
            rest = rest[10:]
        elif rest.startswith("0") and len(rest) >= 9:
            # 室話：02 區碼 10 碼，其他 9 碼
            take = 10 if rest[:2] == "02" and len(rest) >= 10 else 9
            result.append(rest[:take])
            rest = rest[take:]
        else:
            result.append(rest)
            break
    if rest and len(rest) >= 7:
        result.append(rest)
    return result if result else [digits]


def normalize_phones(v) -> list[str]:
    """傳回乾淨的電話列表（保留所有電話，第一支為主）。"""
    if v is None or v == "":
        return []
    s = str(v)
    out = []
    for piece in PHONE_SEP.split(s):
        digits = re.sub(r"\D", "", piece)
        if not digits:
            continue
        # 國際碼 886 開頭 → 去掉換成 0 開頭（886952172357 → 0952172357）
        if digits.startswith("886") and len(digits) >= 11:
            digits = "0" + digits[3:]
        # 手機少前導 0：9 開頭 9 碼 → 補 0
        if len(digits) == 9 and digits.startswith("9"):
            digits = "0" + digits
        # 市話少前導 0：8~9 碼且 4/5/6/7 開頭（台灣 04/05/06/07 區碼）→ 補 0
        # 例：48975018 → 048975018; 422610890 → 0422610890
        elif len(digits) in (8, 9) and digits[0] in "4567":
            digits = "0" + digits
        # 太短跳過（< 7 碼）
        if len(digits) < 7:
            continue
        # >= 12 碼可能是兩支電話沒分開 → 嘗試拆
        if len(digits) >= 12:
            for d in _split_concatenated_phones(digits):
                if len(d) >= 7 and d not in out:
                    out.append(d)
        else:
            if digits not in out:
                out.append(digits)
    return out


def clean_address(v) -> list[str]:
    """傳回去 zip prefix、拆 \\n 後的地址列表。"""
    if v is None or v == "":
        return []
    s = str(v).strip()
    addrs = []
    for line in re.split(r"[\n、]", s):
        line = line.strip()
        if not line:
            continue
        # 去 3 碼郵遞區號開頭（5xx, 4xx, 6xx, ...）
        line = re.sub(r"^\d{3}\s*", "", line)
        if line:
            addrs.append(line)
    return addrs


def parse_amounts(v) -> list[tuple[float, int]]:
    """'1600' → [(1600,1)]; '1600*2' → [(1600,2)]; '2500+1600' → [(2500,1),(1600,1)]"""
    if v is None or v == "":
        return []
    if isinstance(v, (int, float)):
        return [(float(v), 1)]
    s = str(v).strip()
    # 拆 + ＋ 換行 、 , — 多筆金額常用換行分隔，必須拆開避免被當成一個超大數字
    parts = re.split(r"\s*[+＋、,]\s*|\s*[\n\r]+\s*", s)
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        m = re.match(r"^(\d+(?:\.\d+)?)\s*[*×xX]\s*(\d+)$", p)
        if m:
            out.append((float(m.group(1)), int(m.group(2))))
            continue
        # 純數字（可能帶逗號）
        n = re.sub(r"[^\d.]", "", p)
        if n:
            try:
                out.append((float(n), 1))
            except ValueError:
                pass
    return out


def parse_brands(v) -> list[str]:
    """'LG+三洋' → ['LG','三洋']; '三洋*2' → ['三洋','三洋']"""
    if v is None or v == "":
        return []
    s = str(v).strip()
    parts = re.split(r"\s*[+＋、,]\s*", s)
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        # 去尾端品牌型號雜訊（保留主品牌）
        # "日立冷RAS-50JB" → 保留全字串作為品牌資訊
        m = re.match(r"^(.+?)\s*[*×xX]\s*(\d+)$", p)
        if m:
            out.extend([m.group(1).strip()] * int(m.group(2)))
        else:
            out.append(p)
    return out


def parse_payment(v) -> str:
    if not v:
        return ""
    s = str(v).strip()
    if "現金" in s:
        return "cash"
    if "轉帳" in s or "匯款" in s:
        return "transfer"
    if "刷卡" in s or "信用卡" in s or "card" in s.lower():
        return "card"
    if "line" in s.lower():
        return "line_pay"
    return ""


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------
HEADERS = [
    "source_file", "sheet", "row_no", "format",
    "date", "name", "phones", "addresses",
    "brand_raw", "amount_raw", "item_raw", "payment", "source", "machine_id",
    # 新欄位 (2026-05-24)：舊系統 csv 帶來的擴充
    "legacy_code",   # 清洗編號（115-EA247 之類）
    "staff_name",    # 服務人員（阿翰、祥瑋、阿辰...）
    "actual_amount", # 實收金額（扣完折扣加價）
]


def extract_file(path: Path, writer):
    print(f"  讀 {path.name}")
    wb = load_workbook(str(path))
    # 從檔名抓年份備用（解 '7月24日' 這種沒年份的日期）
    m = re.match(r"^(\d{4})", path.name)
    fallback_year = int(m.group(1)) if m else None

    n_rows = 0
    for sh in wb.worksheets:
        fmt, data_start = detect_format(sh)
        cols = COL_MAPS[fmt]

        for r in range(data_start, sh.max_row + 1):
            def get(col_key):
                idx = cols.get(col_key)
                if idx is None:
                    return None
                # idx 是 0-based，openpyxl 是 1-based
                if idx + 1 > sh.max_column:
                    return None
                return sh.cell(r, idx + 1).value

            date_val = get("date")
            name_val = get("name")
            phone_val = get("phone")
            addr_val = get("addr")
            brand_val = get("brand")
            amount_val = get("amount")
            item_val = get("item")
            payment_val = get("payment")
            source_val = get("source")
            machine_id_val = get("machine_id")

            # 整列空白跳過
            row_vals = [date_val, name_val, phone_val, addr_val, brand_val,
                        amount_val, item_val]
            if all(v in (None, "", 0) for v in row_vals):
                continue

            date_str = to_date_str(date_val, fallback_year)
            phones = normalize_phones(phone_val)
            addrs = clean_address(addr_val)

            writer.writerow({
                "source_file": path.name,
                "sheet": sh.title,
                "row_no": r,
                "format": fmt,
                "date": date_str,
                "name": (str(name_val).strip() if name_val else ""),
                "phones": "|".join(phones),
                "addresses": "|".join(addrs),
                "brand_raw": (str(brand_val).strip() if brand_val else ""),
                "amount_raw": (str(amount_val) if amount_val not in (None, "") else ""),
                "item_raw": (str(item_val).strip() if item_val else ""),
                "payment": parse_payment(payment_val),
                "source": (str(source_val).strip() if source_val else ""),
                "machine_id": (str(machine_id_val).strip() if machine_id_val else ""),
                "legacy_code": "",
                "staff_name": "",
                "actual_amount": "",
            })
            n_rows += 1
    print(f"    → {n_rows} rows")


def main():
    files = sorted(p for p in CLEANED_DIR.glob("*_cleaned.xlsx") if not p.name.startswith("~$"))
    if not files:
        print(f"找不到 cleaned 檔案：{CLEANED_DIR}")
        return

    print(f"=== Step 1: Extract ({len(files)} files) ===\n")
    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        writer.writeheader()
        for p in files:
            extract_file(p, writer)

    print(f"\n寫出：{OUT_CSV}")


if __name__ == "__main__":
    main()
