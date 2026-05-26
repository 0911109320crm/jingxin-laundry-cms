"""
把 out/*.csv 切成兩批：

  out/clean/         ✅ 可信客戶（10,891 位）相關完整資料
                     直接給 import-legacy.mjs --clean 匯入

  out/customers_review.xlsx
                     ⚠ 三類問題客戶（電話可疑/地址可疑/無訂單）
                     讓老闆娘逐筆勾選後再決定哪些要匯入

分類規則沿用 customers_overview.py 的 classify_customer()
"""
from __future__ import annotations
import csv
import re
import shutil
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "out"
CLEAN = OUT / "clean"
REVIEW = OUT / "customers_review.xlsx"


def load(name):
    with open(OUT / name, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


print("讀 csv...")
customers = load("customers.csv")
phones = load("customer_phones.csv")
addresses = load("customer_addresses.csv")
machines = load("machines.csv")
orders = load("orders.csv")
order_items = load("order_items.csv")
print(f"  原始：{len(customers)} 客戶 / {len(phones)} 電話 / "
      f"{len(addresses)} 地址 / {len(machines)} 機器 / "
      f"{len(orders)} 訂單 / {len(order_items)} 品項")

phones_by_cust = defaultdict(list)
for p in phones:
    phones_by_cust[p["customer_id"]].append(p)

addrs_by_cust = defaultdict(list)
for a in addresses:
    addrs_by_cust[a["customer_id"]].append(a)

machines_by_cust = defaultdict(list)
for m in machines:
    machines_by_cust[m["customer_id"]].append(m)

orders_by_cust = defaultdict(list)
for o in orders:
    orders_by_cust[o["customer_id"]].append(o)


def classify_customer(c, ps, a_list, n_orders):
    """傳回 ('ok'|'phone'|'address'|'no_order', 原因)"""
    bad_phone_reasons = []
    valid_phone_count = 0
    for p in ps:
        d = re.sub(r"\D", "", p["phone"])
        if len(d) < 9:
            bad_phone_reasons.append(f"{p['phone']} 只 {len(d)} 碼")
            continue
        if len(d) > 11:
            bad_phone_reasons.append(f"{p['phone']} 有 {len(d)} 碼")
            continue
        if len(set(d)) <= 2:
            bad_phone_reasons.append(f"{p['phone']} 數字單調")
            continue
        if d == "1234567890":
            bad_phone_reasons.append("測試假電話")
            continue
        valid_phone_count += 1
    if valid_phone_count == 0:
        return ("phone", "；".join(bad_phone_reasons) or "無有效電話")

    bad_addr_reasons = []
    valid_addr_count = 0
    for a in a_list:
        addr = a["address"]
        if a["county"] == "未分類" and a["district"] == "未分類":
            bad_addr_reasons.append(f"{addr} 縣市鄉鎮全未解析")
            continue
        if not re.search(r"號|樓|室|巷|弄|段|村|里", addr):
            bad_addr_reasons.append(f"{addr} 無號樓巷弄")
            continue
        valid_addr_count += 1
    if a_list and valid_addr_count == 0:
        return ("address", "；".join(bad_addr_reasons[:3]) or "地址無效")

    if n_orders == 0:
        return ("no_order", "只有資料、無服務紀錄")

    return ("ok", "")


# ── 分類 ─────────────────────────────────────────────────
print("\n分類客戶...")
buckets: dict[str, list] = {"ok": [], "phone": [], "address": [], "no_order": []}
cust_bucket: dict[str, str] = {}  # customer_id → bucket key

for c in customers:
    cid = c["id"]
    ps = phones_by_cust[cid]
    a_list = addrs_by_cust[cid]
    n_orders = len(orders_by_cust[cid])
    cat, reason = classify_customer(c, ps, a_list, n_orders)
    buckets[cat].append((c, reason))
    cust_bucket[cid] = cat

for k in ["ok", "phone", "address", "no_order"]:
    print(f"  {k}: {len(buckets[k])} 位")

ok_ids = {c["id"] for c, _ in buckets["ok"]}


# ── 切 CSV → out/clean/ ──────────────────────────────────
print(f"\n寫入 clean csv → {CLEAN}")
CLEAN.mkdir(exist_ok=True)


def write_filtered_csv(name, rows, keep_pred):
    kept = [r for r in rows if keep_pred(r)]
    if not rows:
        return 0
    fields = list(rows[0].keys())
    with open(CLEAN / name, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(kept)
    return len(kept)


def is_ok_cust(r):
    return r.get("customer_id", r.get("id")) in ok_ids or r.get("id") in ok_ids

n_c = write_filtered_csv("customers.csv", customers, lambda r: r["id"] in ok_ids)
n_p = write_filtered_csv("customer_phones.csv", phones, lambda r: r["customer_id"] in ok_ids)
n_a = write_filtered_csv("customer_addresses.csv", addresses, lambda r: r["customer_id"] in ok_ids)
n_m = write_filtered_csv("machines.csv", machines, lambda r: r["customer_id"] in ok_ids)

# orders 依 customer_id 過濾，order_items 依保留的 order_id 過濾
ok_orders = [o for o in orders if o["customer_id"] in ok_ids]
ok_order_ids = {o["id"] for o in ok_orders}
with open(CLEAN / "orders.csv", "w", encoding="utf-8", newline="") as f:
    w = csv.DictWriter(f, fieldnames=list(orders[0].keys()))
    w.writeheader()
    w.writerows(ok_orders)
n_o = len(ok_orders)

ok_items = [it for it in order_items if it["order_id"] in ok_order_ids]
with open(CLEAN / "order_items.csv", "w", encoding="utf-8", newline="") as f:
    w = csv.DictWriter(f, fieldnames=list(order_items[0].keys()))
    w.writeheader()
    w.writerows(ok_items)
n_i = len(ok_items)

print(f"  customers.csv:           {n_c}")
print(f"  customer_phones.csv:     {n_p}")
print(f"  customer_addresses.csv:  {n_a}")
print(f"  machines.csv:            {n_m}")
print(f"  orders.csv:              {n_o}")
print(f"  order_items.csv:         {n_i}")


# ── 寫 review xlsx ──────────────────────────────────────
print(f"\n產出待核對 xlsx → {REVIEW}")
MTYPE_LABEL = {"washing_machine": "洗", "air_conditioner": "冷",
               "mattress": "床", "sofa": "沙", "other": "其"}


def row_for_review(c, reason):
    cid = c["id"]
    ps = sorted(phones_by_cust[cid],
                key=lambda x: (0 if x["is_primary"] == "true" else 1,
                               int(x.get("sort_order") or 0)))
    extra_phones = [p["phone"] + (f"({p['label']})" if p["label"] else "")
                    for p in ps if p["is_primary"] != "true"]
    a_list = addrs_by_cust[cid]
    default_addr = next((a for a in a_list if a["is_default"] == "true"),
                        a_list[0] if a_list else None)
    default_str = (
        f"{default_addr['county']}{default_addr['district']} {default_addr['address']}"
        if default_addr else ""
    )
    other_addrs = [
        f"{a['county']}{a['district']} {a['address']}"
        for a in a_list if a is not default_addr
    ]
    m_list = machines_by_cust[cid]
    type_count = defaultdict(int)
    for m in m_list:
        type_count[m["type"]] += 1
    type_summary = " ".join(f"{MTYPE_LABEL.get(t, t)}×{n}" for t, n in type_count.items())
    machine_details = "; ".join(
        f"{MTYPE_LABEL.get(m['type'], m['type'])} {m['brand'] or '(無)'}"
        for m in m_list
    )
    o_list = sorted(orders_by_cust[cid], key=lambda x: x["service_at"])
    n_orders = len(o_list)
    first_date = o_list[0]["service_at"][:10] if o_list else ""
    last_date = o_list[-1]["service_at"][:10] if o_list else ""
    total_spent = sum(float(o["total"]) for o in o_list if o["total"])

    return [
        "",  # 老闆娘勾選欄（留空，她填 v / x）
        c["code"], c["name"], c["phone"],
        ", ".join(extra_phones),
        default_str,
        "; ".join(other_addrs),
        type_summary,
        machine_details[:150],
        n_orders, first_date, last_date,
        f"{total_spent:,.0f}" if total_spent else "",
        reason,
    ]


HEADERS = [
    "決定 (v=匯入 / x=不要 / 留空=待定)",
    "編號", "姓名", "主電話", "副電話", "預設地址", "其他地址",
    "機器(數量)", "機器詳細", "訂單數", "首次服務", "末次服務",
    "累計金額", "問題原因",
]
WIDTHS = [28, 12, 12, 14, 28, 32, 30, 14, 30, 8, 12, 12, 11, 40]

wb = Workbook()
wb.remove(wb.active)

# 摘要 sheet
sum_sh = wb.create_sheet("摘要")
sum_sh.append(["分類", "客戶數", "原因 / 說明"])
sum_sh.append(["⚠ 電話可疑", len(buckets["phone"]),
               "電話 < 9 碼、> 11 碼、單調數字（如 00000000、1234567890）"])
sum_sh.append(["⚠ 地址可疑", len(buckets["address"]),
               "地址縣市鄉鎮無法解析，或無「號/樓/巷/弄/段/村/里」"])
sum_sh.append(["⚠ 無訂單紀錄", len(buckets["no_order"]),
               "客戶資料齊但沒實際服務紀錄（多為 Google 表單預約沒接到）"])
sum_sh.append([])
sum_sh.append(["總計待核對", sum(len(buckets[k]) for k in ["phone", "address", "no_order"]), ""])
sum_sh.append([])
sum_sh.append(["說明:"])
sum_sh.append(["  在每一張 sheet 的第一欄填入 v（要匯入）或 x（不要）"])
sum_sh.append(["  完成後存檔回傳，後續會用此檔再次批次匯入"])

bold = Font(bold=True)
header_fill = PatternFill("solid", fgColor="FFF3CD")
sum_fill = PatternFill("solid", fgColor="EEEEEE")
for c in range(1, 4):
    sum_sh.cell(1, c).font = bold
    sum_sh.cell(1, c).fill = sum_fill
sum_sh.column_dimensions["A"].width = 25
sum_sh.column_dimensions["B"].width = 12
sum_sh.column_dimensions["C"].width = 70

SHEET_DEFS = [
    ("⚠ 電話可疑", "phone"),
    ("⚠ 地址可疑", "address"),
    ("⚠ 無訂單紀錄", "no_order"),
]

for sheet_name, bk in SHEET_DEFS:
    sh = wb.create_sheet(sheet_name)
    sh.append(HEADERS)
    for c, reason in buckets[bk]:
        sh.append(row_for_review(c, reason))
    for col in range(1, len(HEADERS) + 1):
        sh.cell(1, col).font = bold
        sh.cell(1, col).fill = header_fill
        sh.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")
    sh.freeze_panes = "C2"
    for i, w in enumerate(WIDTHS, start=1):
        sh.column_dimensions[get_column_letter(i)].width = w
    sh.auto_filter.ref = sh.dimensions

wb.save(str(REVIEW))
print(f"  {REVIEW}")
print(f"\n完成。下一步：")
print(f"  1. 老闆娘看 {REVIEW.name} 決定要不要匯入那 126 筆問題客戶")
print(f"  2. 用 import-legacy.mjs --clean 把 {CLEAN}/ 的 10,891 位先灌進 Supabase")
