"""
讀老闆娘在 merged_all.xlsx「未識別備註」sheet 填好的審核結果，
產出 review_decisions.json 給 build.py 套用。

執行:
  python apply_review.py

輸出:
  out/review_decisions.json — 以 (來源檔, 工作表, 原列號) 為 key 的決定列表

build.py 之後讀這個 json，對應每筆 rows.csv 套用：
  1. 忽略           → 不做事
  2. 客戶備註       → customer.note 追加
  3. 訂單備註       → order.note 追加
  4. 機器備註       → machine.note 追加
  5. 地址補充       → address.label 追加
  6. 副電話         → customer_phones 新增一筆
  7. 店家名稱       → customer.note 加 "[商家] xxx"
  8. 加價           → order_adjustments 新增 addon
  9. 折扣           → order_adjustments 新增 discount
 10. 訂單取消       → order.status = cancelled
 11. 機器型號       → machine.model
"""
from __future__ import annotations
import json
import re
from pathlib import Path
from openpyxl import load_workbook

MERGED = Path(__file__).parent / "out" / "merged_all.xlsx"
OUT = Path(__file__).parent / "out" / "review_decisions.json"

# 處理方式 → 動作代號
DECISION_MAP = {
    "1": "ignore",
    "2": "customer_note",
    "3": "order_note",
    "4": "machine_note",
    "5": "address_label",
    "6": "secondary_phone",
    "7": "business_name",
    "8": "addon",
    "9": "discount",
    "10": "cancel",
    "11": "machine_model",
}


def parse_decision(s: str) -> str | None:
    """從『8. 加價 - 收額外費用（要填金額）』抓出 '8'，回 'addon'。"""
    if not s:
        return None
    s = str(s).strip()
    m = re.match(r"^(\d+)\.", s)
    if not m:
        return None
    return DECISION_MAP.get(m.group(1))


def main():
    print(f"讀 {MERGED}")
    wb = load_workbook(str(MERGED))
    if "未識別備註" not in wb.sheetnames:
        print("✗ 找不到「未識別備註」工作表")
        return
    sh = wb["未識別備註"]

    decisions = []
    counts = {"total": 0, "ignored": 0, "blank": 0, "invalid": 0}

    # 欄位順序：A=來源檔 B=工作表 C=原列號 D=日期 E=姓名 F=電話 G=備註原文 H=可能原因
    #          I=處理方式 J=補充說明 K=金額
    for row in sh.iter_rows(min_row=2, values_only=True):
        if not row or all(v in (None, "") for v in row):
            continue
        counts["total"] += 1
        src_file, sheet, row_no = row[0], row[1], row[2]
        note_raw = row[6] or ""
        decision_raw = row[8]
        extra = row[9] or ""
        amount = row[10]

        action = parse_decision(decision_raw)
        if action is None:
            if decision_raw in (None, ""):
                counts["blank"] += 1
                action = "ignore"  # 沒填預設忽略
            else:
                counts["invalid"] += 1
                print(f"  ⚠ {src_file} {sheet} row{row_no}: 無效選項 {decision_raw!r}")
                continue
        if action == "ignore":
            counts["ignored"] += 1
            continue

        # 加價/折扣需要金額
        if action in ("addon", "discount"):
            try:
                amount = float(amount)
            except (TypeError, ValueError):
                print(f"  ⚠ {src_file} {sheet} row{row_no}: {action} 但金額空白或無效 → 跳過")
                continue

        decisions.append({
            "source_file": src_file,
            "sheet": sheet,
            "row_no": row_no,
            "action": action,
            "note_raw": note_raw,
            "extra": extra,
            "amount": amount if action in ("addon", "discount") else None,
        })

    OUT.write_text(json.dumps(decisions, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n寫出 {OUT}")
    print(f"\n統計:")
    print(f"  總列數: {counts['total']}")
    print(f"  空白未填: {counts['blank']} (視為忽略)")
    print(f"  選忽略:   {counts['ignored']}")
    print(f"  無效選項: {counts['invalid']}")
    print(f"  實際採用: {len(decisions)}")
    # 按 action 統計
    by_action = {}
    for d in decisions:
        by_action[d["action"]] = by_action.get(d["action"], 0) + 1
    print(f"\n各動作分佈:")
    for a, n in sorted(by_action.items(), key=lambda x: -x[1]):
        print(f"  {a:20s} {n}")


if __name__ == "__main__":
    main()
