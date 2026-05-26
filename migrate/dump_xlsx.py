"""把 customers_overview / manual_review 轉成可讀預覽"""
import openpyxl

def dump(path, max_rows=20):
    print(f'\n{"="*70}\n{path}\n{"="*70}')
    wb = openpyxl.load_workbook(path, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f'\n--- Sheet: {sn}  ({ws.max_row} rows × {ws.max_column} cols) ---')
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                print(f'  ... (還有 {ws.max_row - max_rows} 筆)')
                break
            cells = [str(c) if c is not None else '' for c in row]
            print('  ' + ' | '.join(cells))

dump('out/customers_overview.xlsx', max_rows=15)
dump('out/manual_review.xlsx', max_rows=20)
