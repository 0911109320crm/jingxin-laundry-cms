"""
把 cleaned/ 8 個 xlsx 統整成單一檔案：merged_all.xlsx

3 種格式：
  A (2016~2021): 編號, 日期, 姓名, 電話, 地址, 廠牌, 金額, 來源處, 卡號, 備註
  B (2022, 2023/2月+): 日期, 姓名, 電話, 地址, 項目, 廠牌, 付款方式, 金額,
                       增加, 折扣, 實收, 服務人員, 機器編號, 備註
  C (2023/1月，無 header): 日期, 姓名, 電話, 地址, 廠牌, 金額, 折扣/備註, 機器編號,
                          服務人員, 標記符號

輸出 sheets:
  1. 全部訂單     ─ 所有 10,508 列，欄位齊全 + 備註自動分類
  2. 分類統計     ─ 各 tag 出現次數
  3. 未識別備註   ─ 沒分到任何類別的備註，含原因
  4. 分類規則說明 ─ 列出 22 個 tag 的辨識規則，方便老闆娘調整
"""
from __future__ import annotations
import re
from datetime import datetime, date
from pathlib import Path
from collections import Counter

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


CLEANED = Path(r"C:\RenStudio\case\washinmachine\顧客名單\cleaned")
OUT = Path(__file__).parent / "out" / "merged_all.xlsx"


# ---------------------------------------------------------------------------
# 格式偵測 + COL_MAP
# ---------------------------------------------------------------------------
COL_MAPS = {
    "A": {
        "row_no": 0, "date": 1, "name": 2, "phone": 3, "addr": 4,
        "brand": 5, "amount": 6, "source": 7, "card_no": 8, "note": 9,
    },
    "B": {
        "date": 0, "name": 1, "phone": 2, "addr": 3, "item": 4,
        "brand": 5, "payment": 6, "amount": 7,
        "addon": 8, "discount": 9, "actual": 10,
        "staff": 11, "machine_code": 12, "note": 13,
    },
    "C": {
        "date": 0, "name": 1, "phone": 2, "addr": 3, "brand": 4,
        "amount": 5, "discount": 6, "machine_code": 7,
        "staff": 8, "mark": 9,
    },
}


def detect_format(sheet) -> tuple[str, int]:
    r1 = [sheet.cell(1, c).value for c in range(1, min(15, sheet.max_column + 1))]
    has_header = any(isinstance(v, str) and v in ("日期", "姓名", "編號") for v in r1)
    has_item_header = any(isinstance(v, str) and "項目" in v for v in r1)
    data_start = 2 if has_header else 1
    # 找第一個有效資料列
    while data_start <= sheet.max_row:
        rd = [sheet.cell(data_start, c).value for c in range(1, min(15, sheet.max_column + 1))]
        if not all(v in (None, "") for v in rd):
            break
        data_start += 1

    rd = [sheet.cell(data_start, c).value for c in range(1, min(15, sheet.max_column + 1))]
    if isinstance(rd[0], int) and 0 < rd[0] < 999:
        return "A", data_start
    if has_item_header:
        return "B", data_start
    return "C", data_start


# ---------------------------------------------------------------------------
# 備註分類規則
# ---------------------------------------------------------------------------
# 每個 tag = (name, label_中文, regex, description)
RULES: list[tuple[str, str, re.Pattern, str]] = [
    ("promo_review",     "促銷-打卡評論",
     re.compile(r"打卡|評論"),
     "含「打卡」或「評論」字樣 → 折扣（GMB 評論回饋）"),

    ("promo_group",      "促銷-團購",
     re.compile(r"團購|團-|團$"),
     "含「團購」或單字「團」 → 團購折扣"),

    ("promo_loyalty",    "促銷-固定清洗",
     re.compile(r"固定清洗"),
     "含「固定清洗」 → 老客戶回購折扣"),

    ("discount_explicit","折扣-明確金額",
     re.compile(r"\(-\d+\)|^折\-?\d+|^折扣|^-\d+$|減\d+|折\d+"),
     "「(-300)」「折100」「-100」「減 200」「折扣」這種純折扣標記"),

    ("addon_dismount",   "加價-拆解費",
     re.compile(r"拆解|^拆\d+"),
     "「拆解費」「拆300」這種拆機需額外收費"),

    ("addon_travel",     "加價-車馬費/搬機",
     re.compile(r"車馬|遠距離|搬機|^太?遠|遠\+?\d+|\+\d+\s*搬|協助搬|搬到\d+樓"),
     "「車馬費」「遠+300」「太遠」「協助搬到 1 樓」 → 遠距離或搬機加價"),

    ("addon_parts",      "加價-耗材/零件",
     re.compile(r"皮帶|避震|橡膠圈|吊桿|壓桿|插管|插\d+|急\+|濾網|螺母|排水管|鐵座|膠條|膠圈|像膠圈|卡盤|換盤|軸心|洗衣盤"),
     "「換皮帶」「橡膠圈」「濾網」「卡盤」「軸心」「洗衣盤」這種維修耗材"),

    ("addon_special_machine", "加價-特殊機型",
     re.compile(r"特殊機型|非服務範圍"),
     "「冷氣特殊機型」「非服務範圍內收原價」 → 加收費用"),

    ("addon_other",      "加價-其他",
     re.compile(r"^\$?\d{2,4}\s*\(.*?(\+|加|增)"),
     "「+1600 元」「加 300」之類的加價字串"),

    ("pandemic_reschedule","狀態-疫情改期",
     re.compile(r"疫改|疫延|疫情"),
     "「疫改 6/26」「疫延 6/23」 → 2021-2022 疫情期間改期"),

    ("status_reschedule","狀態-改期",
     re.compile(r"改期|延後|順路改"),
     "「改期」「順路改期」「取消延後」 → 一般改期"),

    ("status_cancel",    "狀態-取消",
     re.compile(r"^取消|放鳥"),
     "「取消」「放鳥」 → 訂單取消"),

    ("status_failed",    "狀態-無法完成",
     re.compile(r"未洗|卡死|拆不起|斷頭|螺絲生?[繡鏽]|軸.*壞|軸芯|無汙、拆|美規無法拆|沒拆起|風鼓用壞|機板.*破壞"),
     "「未洗」「卡死沒洗」「拆不起來」「軸心壞了」「美規無法拆」「風鼓用壞」 → 到場無法服務"),

    ("price_only",       "金額補充",
     re.compile(r"^\d{3,5}(\s*[\+xX×*\-、，]\s*\d+)*\s*元?$|^\d{3,5}\s*\n\s*\d{3,5}|^[、，]?\s*\d{3,5}\s*元?$|^[一-鿿]{0,3}\$?\d{3,5}\s*元?$|加\d+\s*元|\d+\s*X\s*\d+\s*元"),
     "「1600」「3800+1600」「1800X2元」「加1600元」「三洋1600」「惠2300元」 → 純金額補充（已含在訂單）"),

    ("service_repair",   "服務-維修",
     re.compile(r"^維修$|^修理$|換避震|更換|^維修\b|疏通|改排水"),
     "「修理」「維修」「換避震」「疏通」「改排水」 → 維修服務（非清洗）"),

    ("service_combo",    "機台-組合（N洗N冷）",
     re.compile(r"\d+\s*洗\s*\d+\s*冷|\d+\s*洗\s*\d+\s*滾|\d+\s*冷\s*\d+\s*洗|\d+\s*滾\s*\d+\s*洗|^共\s*\d+|^\d+洗$|^\d+冷$|^\d+滾$"),
     "「1洗1冷」「共2洗1冷」「2洗」「3冷」 → 機型組合（多項服務）"),

    ("machine_count",    "機台-數量補充",
     re.compile(r"^[一二三四五六七八九十兩]\s*台|^\d+\s*台|外加\d+|家有\d+台|家有[一二三四五六七八九十兩]+台|\+\d+\s*台|每台"),
     "「二台」「兩台」「家有 2 台」「每台減 200」 → 該客有多台需服務"),

    ("machine_model",    "機台-型號代碼",
     re.compile(r"^[A-Za-z][A-Za-z0-9\-/.]{3,}$|(?:[A-Z]{2,}[\-]?\d{2,}[A-Z0-9\-]*)|(?:[冷滾洗膠日東國三大聲LG][A-Za-z0-9\-]{4,})"),
     "「WD-16NEB」「日立SF-BD2200T」「冷RAS-50NK」「滾NA-V178BN」 → 機器型號（含中文前綴）"),

    ("old_system_code",  "舊系統客戶編號",
     re.compile(r"舊\s*\d{2,4}\s*[-/]\s*\d{1,2}\s*[-/]\s*\d{1,2}"),
     "「舊 106-2-22」 → 老闆娘舊系統的客戶編號（重要！）"),

    ("paper_form",       "紙本表編號",
     re.compile(r"^表\s*\d+"),
     "「表 686」 → 紙本訂單表編號"),

    ("secondary_phone",  "副電話",
     re.compile(r"^\D*0?\d{8,11}\D*$|^\d{1,2}\-\d{6,8}$|0\d-\d{6,8}"),
     "9 碼以上或含市話 pattern（0X-XXXXXXX） → 副電話"),

    ("business_name",    "店家/機關",
     re.compile(r"(店|館|廟|宮|堂|寺|別墅|別野|別院|消防局|警察局?|派出所|衛生所|銀行|郵局|學校|國小|國中|高中|大學|診所|眼科|醫院|事務所|工坊|工廠|食堂|餐廳|超市|超商|大樓|社區|公司|診所|小吃|醬油|攤|麵|戲|樓)$|^[一-鿿]{2,4}小吃|^[一-鿿]{2,4}醬|布袋戲|警察局|消防局"),
     "結尾為機關/店家名（如「建忠眼科」「呷味先小吃」「天星別野」「義春園布袋戲」「3冷（田尾消防局）」） → 客戶為店家"),

    ("machine_location", "機器位置/路線",
     re.compile(r"洗衣機在|機器在|機台在|住家在|地址.*：|地址是|在\d+樓|在.{1,3}樓|樓梯|電梯|可導航|要導|導航|頂樓|客廳|廚房|浴室|廁所|陽台|後院|裡面|外面|巷子|路.{0,5}停車|停車格|轉角"),
     "「洗衣機在浴室」「住家在巷子裡」「導航到太極新村」「樓梯很小」 → 機器位置 / 找路說明"),

    ("machine_condition","機器狀況",
     re.compile(r"很髒|很臭|霧|生繡|生鏽|聲音|大聲|破壞|沒拆|問題|髒|臭|壞|機板|風鼓|軸恥|有.{1,3}情況|15年沒洗|多年沒洗|第一次洗"),
     "「軸心大聲」「機板螺絲已破壞」「桶子有白霧情況」 → 機器狀態 / 風險預告"),

    ("customer_note",    "客戶特殊狀況",
     re.compile(r"聽障|要收據|常客|老顧客|每年|定期清|定洗"),
     "「聽障」「要收據」「老顧客」「每年都清洗」「定期」 → 客戶屬性或習慣"),

    ("social_fan",       "FB 粉絲/部落客",
     re.compile(r"^粉[:：]|部落客|^生@|粉絲"),
     "「粉:olive hsu」「(部落客)」「生@WINNI」 → 來源為 FB 粉絲頁的客戶"),

    ("machine_capacity", "機台容量/規格",
     re.compile(r"\d+\s*kg|\d+\s*公斤|^\d+\s*[大小]台|大台|小台|頂級|高階|\d+以上機型|18\s*以上"),
     "「11kg」「大台國際18kg」「洗衣機18以上」「63以上機型」 → 機台容量規格（影響定價）"),

    ("relationship",     "關係/介紹來源",
     re.compile(r"親友|鄰居|朋友|同事|堂|嫂|月嫂|姊夫|姊妹|姊朋友|姐姐|妹妹|哥哥|弟弟|阿姨|舅|姑|叔|伯|公|婆|爸|媽|娘家|外甥|姪|.{1,3}的朋友"),
     "「堂嫂」「鄰居朋友」「權鈺的朋友」「娘家」 → 介紹關係或客戶關係"),

    ("landmark",         "地址補充-地標",
     re.compile(r"隔壁|對面|^旁|前面|後面|附近|^.{2,4}(旁|對面)"),
     "「福寧宮隔壁」「虎尾高中對面」 → 找路標"),

    ("payment_transfer", "付款-匯款",
     re.compile(r"^匯款$|轉帳|ATM"),
     "「匯款」「轉帳」 → 付款方式為匯款"),

    ("loyalty_discount", "舊客戶優惠",
     re.compile(r"舊客.*不加|每台減|定期|固定洗|活動價"),
     "「舊客不加300」「每台減200」「定期」「活動價」 → 老客戶優惠"),
]


def classify(note: str) -> tuple[list[str], list[str]]:
    """傳回 (matched_tags, reasons_for_unmatched)"""
    if not note:
        return [], ["備註為空"]
    s = note.strip()
    if not s:
        return [], ["備註為空白字元"]
    tags = []
    for tag, _, pat, _ in RULES:
        if pat.search(s):
            tags.append(tag)
    if not tags:
        reasons = []
        if len(s) <= 3 and not any(c.isdigit() for c in s):
            reasons.append(f"太短 ({len(s)} 字) 且無數字，無法套用 pattern")
        if re.fullmatch(r"[\w一-鿿]{1,4}", s):
            reasons.append("看起來是縮寫或代號（如『下國4』『放鳥』）")
        if "\n" in s:
            reasons.append("多行內容，混合多種類型")
        if not reasons:
            reasons.append("未匹配任何 pattern — 可能是新型描述")
        return [], reasons
    return tags, []


# ---------------------------------------------------------------------------
# 讀檔
# ---------------------------------------------------------------------------
def to_date_str(v) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        try:
            dt = datetime(1899, 12, 30) + __import__("datetime").timedelta(days=int(v))
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return ""
    if isinstance(v, str):
        s = v.strip()
        m = re.match(r"^(\d{1,2})月(\d{1,2})日?$", s)
        if m:
            return s  # 保留原文，年份從檔名抓
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return s
    return str(v)


def cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    return str(v)


def extract_rows(fp: Path) -> list[dict]:
    print(f"  讀 {fp.name}")
    wb = load_workbook(str(fp), read_only=True)
    rows = []
    for sh in wb.worksheets:
        fmt, data_start = detect_format(sh)
        cols = COL_MAPS[fmt]

        for r_idx, row in enumerate(
            sh.iter_rows(min_row=data_start, values_only=True),
            start=data_start,
        ):
            # 整列空白跳過
            if not row or all(v in (None, "") for v in row):
                continue

            def g(key):
                idx = cols.get(key)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            note_raw = cell_str(g("note"))
            # Format C 沒有 note 欄，把可能的字串集合
            if fmt == "C":
                note_raw = ""

            row_dict = {
                "來源檔": fp.name.replace("_cleaned.xlsx", ""),
                "工作表": sh.title,
                "原列號": r_idx,
                "格式": fmt,
                "日期": to_date_str(g("date")),
                "姓名": cell_str(g("name")),
                "電話": cell_str(g("phone")),
                "地址": cell_str(g("addr")),
                "項目": cell_str(g("item")),
                "廠牌": cell_str(g("brand")),
                "金額": cell_str(g("amount")),
                "付款方式": cell_str(g("payment")),
                "增加項": cell_str(g("addon")),
                "折扣項": cell_str(g("discount")),
                "實收": cell_str(g("actual")),
                "來源處": cell_str(g("source")),
                "卡號": cell_str(g("card_no")),
                "機器編號": cell_str(g("machine_code")),
                "服務人員": cell_str(g("staff")),
                "標記": cell_str(g("mark")),
                "備註原文": note_raw,
            }
            rows.append(row_dict)
    return rows


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------
def main():
    print("=== 整合所有 cleaned 檔 ===\n")
    all_rows = []
    for fp in sorted(CLEANED.glob("*_cleaned.xlsx")):
        if fp.name.startswith("~$"):
            continue
        all_rows.extend(extract_rows(fp))
    print(f"\n總共 {len(all_rows)} 列\n")

    # ---- 分類備註 ----
    # 同時也分類「增加項」「折扣項」「卡號」（它們也含類似格式的字串）
    tag_counter = Counter()
    unclassified = []  # rows with non-empty note but no tags
    for r in all_rows:
        # 主要看備註，再加上增加項/折扣項一起當「備註整體」
        combined = "\n".join(
            x for x in (r["備註原文"], r["增加項"], r["折扣項"], r["卡號"])
            if x
        )
        tags, reasons = classify(combined)
        r["分類標籤"] = "、".join(tags)
        r["分類_label"] = "、".join(
            label for tag, label, _, _ in RULES if tag in tags
        )
        for t in tags:
            tag_counter[t] += 1
        if combined and not tags:
            unclassified.append((r, reasons))
    print(f"已分類: {sum(tag_counter.values())} 個 tag 命中")
    print(f"未識別備註: {len(unclassified)}")

    # ---- 寫出 xlsx ----
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 0: 老闆娘審核 SOP（放第一張，一開檔就看到）
    sh0 = wb.create_sheet("老闆娘審核 SOP")
    sop_lines = [
        ["📋 未識別備註審核流程"],
        [""],
        ["這份檔案有 5 張工作表，你只需要審核「未識別備註」這張。"],
        [""],
        ["----------------------------------------"],
        ["步驟 1：打開「未識別備註」工作表"],
        ["----------------------------------------"],
        ["共 36 筆，每一筆都是程式無法自動歸類的備註。"],
        [""],
        ["每一列你會看到："],
        ["  G 欄『備註原文』─ 老資料裡寫的字"],
        ["  H 欄『可能原因』─ 為什麼程式抓不到（太短、是縮寫等）"],
        ["  I 欄『處理方式』─ ★你要填的★ 點旁邊三角形，從 11 選 1"],
        ["  J 欄『補充說明』─ 選填，你想多寫什麼都可以"],
        ["  K 欄『金額』─ 只有選『加價』『折扣』時才需要填數字"],
        [""],
        ["----------------------------------------"],
        ["步驟 2：對每一筆從下拉選一個處理方式"],
        ["----------------------------------------"],
        ["11 個選項的意思："],
        [""],
        ["  1. 忽略 - 沒用的資訊       → 直接丟掉，不進系統"],
        ["  2. 客戶備註 - 加進客戶頁   → 例：『聽障』『要收據』『老顧客』"],
        ["  3. 訂單備註 - 這張訂單     → 例：『改洗冷氣』『上下洗只洗上』"],
        ["  4. 機器備註 - 那台機器     → 例：『軸心大聲』『風鼓用壞』"],
        ["  5. 地址補充 - 找路資訊     → 例：『隔壁』『對面』『裡面』"],
        ["  6. 副電話 - 含電話號碼     → 例：『04-22733645』『老公 09xxx』"],
        ["  7. 店家名稱 - 客戶是商家   → 例：『領帶城』『良原醬油』『梅林營區』"],
        ["  8. 加價 - 收額外費用       → 例：『+300 遠』 → 金額欄填 300"],
        ["  9. 折扣 - 給客戶折扣       → 例：『-100』 → 金額欄填 100"],
        ["  10. 訂單取消 - 沒完成      → 例：『美規無法拆』『大同沒拆起』"],
        ["  11. 機器型號               → 例：『WD-S18VW』『AM-G36L』"],
        [""],
        ["----------------------------------------"],
        ["步驟 3：存檔 → 告訴 RC「我審核好了」"],
        ["----------------------------------------"],
        ["RC 會跑一行程式，把你的選擇套用進匯入流程，"],
        ["這 36 筆就會跟其他 10,472 筆一起灌進系統。"],
        [""],
        ["----------------------------------------"],
        ["💡 小撇步"],
        ["----------------------------------------"],
        ["• 真的不確定的，選『1. 忽略』就好，後面在系統裡再補也行"],
        ["• 同一個客戶可能有多筆，可以下拉複製到下一格"],
        ["• 沒填的會被視為『1. 忽略』"],
        ["• 已分類好的 1,683 筆備註不用你看，系統會自動處理"],
        [""],
    ]
    for line in sop_lines:
        sh0.append(line)
    # 樣式
    sh0.column_dimensions["A"].width = 80
    sh0.cell(1, 1).font = Font(bold=True, size=16, color="0066CC")
    for r_idx in range(1, len(sop_lines) + 1):
        cell = sh0.cell(r_idx, 1)
        text = cell.value or ""
        if text.startswith("步驟") or text.startswith("📋") or text.startswith("💡"):
            cell.font = Font(bold=True, size=12, color="0066CC")
        elif text.startswith("--"):
            cell.font = Font(color="999999")
        cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Sheet 1: 全部訂單
    sh1 = wb.create_sheet("全部訂單")
    headers = list(all_rows[0].keys()) if all_rows else []
    sh1.append(headers)
    for r in all_rows:
        sh1.append([r.get(h, "") for h in headers])
    style_sheet(sh1, len(headers))

    # Sheet 2: 分類統計
    sh2 = wb.create_sheet("分類統計")
    sh2.append(["排序", "標籤代號", "標籤中文", "出現次數", "辨識規則說明"])
    rank = 1
    for tag, label, _, desc in RULES:
        n = tag_counter.get(tag, 0)
        sh2.append([rank, tag, label, n, desc])
        rank += 1
    # 把無分類的也加進去
    sh2.append([rank, "(未分類)", "未識別備註", len(unclassified),
                "備註不為空，但任何 pattern 都沒匹配 → 另存『未識別備註』sheet"])
    style_sheet(sh2, 5)
    sh2.column_dimensions["E"].width = 60
    sh2.column_dimensions["C"].width = 18

    # Sheet 3: 未識別備註（含老闆娘審核欄位）
    sh3 = wb.create_sheet("未識別備註")
    sh3.append([
        "來源檔", "工作表", "原列號", "日期", "姓名", "電話",
        "備註原文（無法分類）", "可能原因",
        "處理方式（下拉選）", "補充說明（選填）", "金額（加價/折扣時填）",
    ])
    # 8 → 11 欄都填資料
    for r, reasons in unclassified:
        sh3.append([
            r["來源檔"], r["工作表"], r["原列號"], r["日期"], r["姓名"], r["電話"],
            "\n".join(x for x in (r["備註原文"], r["增加項"], r["折扣項"], r["卡號"]) if x),
            "；".join(reasons),
            "", "", "",  # 給老闆娘填
        ])
    style_sheet(sh3, 11)
    sh3.column_dimensions["G"].width = 40   # 備註原文
    sh3.column_dimensions["H"].width = 30   # 可能原因
    sh3.column_dimensions["I"].width = 26   # 處理方式
    sh3.column_dimensions["J"].width = 24   # 補充說明
    sh3.column_dimensions["K"].width = 12   # 金額

    # 「處理方式」加下拉驗證 — 11 個選項
    DECISIONS = [
        "1. 忽略 - 沒用的資訊",
        "2. 客戶備註 - 加進客戶頁",
        "3. 訂單備註 - 加進這張訂單",
        "4. 機器備註 - 跟那台機器有關",
        "5. 地址補充 - 找路資訊",
        "6. 副電話 - 含電話號碼",
        "7. 店家名稱 - 客戶是商家",
        "8. 加價 - 收額外費用（要填金額）",
        "9. 折扣 - 給客戶折扣（要填金額）",
        "10. 訂單取消 - 沒完成",
        "11. 機器型號",
    ]
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(DECISIONS) + '"',
        allow_blank=True,
        showDropDown=False,  # False = 顯示下拉箭頭
    )
    dv.error = "請從下拉選項中選一個"
    dv.errorTitle = "選錯了"
    dv.prompt = "點旁邊箭頭，選 1~11 其中一個"
    dv.promptTitle = "如何處理這筆？"
    sh3.add_data_validation(dv)
    # 套用到 I 欄（處理方式），從 row 2 開始
    last_row = len(unclassified) + 1
    dv.add(f"I2:I{last_row}")
    sh3.column_dimensions["G"].width = 40
    sh3.column_dimensions["H"].width = 35

    # Sheet 4: 分類規則說明
    sh4 = wb.create_sheet("分類規則說明")
    sh4.append([
        "標籤代號", "標籤中文", "辨識條件 (regex)", "用途說明",
        "建議對應 CMS 欄位",
    ])
    cms_map = {
        "promo_review": "order_adjustments (discount)",
        "promo_group": "order_adjustments (discount)",
        "promo_loyalty": "order_adjustments (discount)",
        "discount_explicit": "order_adjustments (discount)",
        "loyalty_discount": "order_adjustments (discount)『舊客優惠』",
        "addon_dismount": "order_adjustments (addon)『拆解費』",
        "addon_travel": "order_adjustments (addon)『遠距離車馬費』",
        "addon_parts": "order_adjustments (addon)『耗材』",
        "addon_special_machine": "order_adjustments (addon)『特殊機型加價』",
        "addon_other": "order_adjustments (addon)",
        "price_only": "（忽略：金額已在訂單金額欄位）",
        "pandemic_reschedule": "order.note + scheduled_at 改期",
        "status_reschedule": "order.note（標記）",
        "status_cancel": "order.status = cancelled",
        "status_failed": "order.status = cancelled + cancellation_reason",
        "service_repair": "新建 service_item『維修』『疏通』",
        "service_combo": "多開 order_items（每個機型 1 個）",
        "machine_count": "需多開 order_items 或 quantity 調整",
        "machine_model": "machine.model",
        "old_system_code": "customer.note（保留老闆娘舊系統編號對照）",
        "paper_form": "customer.note（紙本表編號）",
        "secondary_phone": "customer_phones（副電話）",
        "business_name": "customer.note（客戶為店家/機關）",
        "relationship": "customer.referrer_id 或 customer.note",
        "landmark": "address.label 或 address.note",
        "machine_location": "machine.note 或 address.label（機器位置）",
        "machine_condition": "machine.note 或 order.note（風險預告）",
        "customer_note": "customer.note（聽障/老顧客等特殊狀況）",
        "social_fan": "customer.source = FB 粉絲團 + customer.note",
        "machine_capacity": "machine.sub_type（容量規格）",
        "payment_transfer": "order.payment_method = transfer",
    }
    for tag, label, pat, desc in RULES:
        sh4.append([tag, label, pat.pattern, desc, cms_map.get(tag, "")])
    style_sheet(sh4, 5)
    sh4.column_dimensions["C"].width = 35
    sh4.column_dimensions["D"].width = 55
    sh4.column_dimensions["E"].width = 35

    wb.save(str(OUT))
    print(f"\n寫出 {OUT}")
    print(f"\n各 sheet:")
    for s in wb.sheetnames:
        print(f"  - {s}: {wb[s].max_row - 1} 列")


def style_sheet(sh, n_cols):
    """套用基本樣式：第一列粗體 + 凍結首列 + 自動欄寬。"""
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="EEEEEE")
    for c in range(1, n_cols + 1):
        cell = sh.cell(1, c)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sh.freeze_panes = "A2"
    for c in range(1, n_cols + 1):
        letter = get_column_letter(c)
        # 預設寬度
        sh.column_dimensions[letter].width = 14


if __name__ == "__main__":
    main()
