"""
讀「淨新Google預約表單.xlsx」轉成跟 rows.csv 同 schema 的 gform_rows.csv。

格式對照：
  gform col → rows.csv 欄位
  0 時間戳記        → date (取日期部分)
  2 姓名            → name
  3 行動電話 + 4 家用電話 → phones (用 | 串接)
  5 清洗地址        → addresses
  6 直立式洗衣機 (廠牌) → 拆成 brand_raw + machine_type=washing_vertical
  7 滾筒式洗衣機 (廠牌)
  9 分離式冷氣 (廠牌)
  11 床墊/沙發 清潔
  8 洗衣機台數     → 影響 quantity
  10 冷氣台數
  12 備註欄         → note_raw
  13 客戶來源       → source

跟 cleaned 重疊判定：用 (主電話, YYYY-MM) 作 key，與 rows.csv 比對，重複跳過。
"""
from __future__ import annotations
import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from extract import normalize_phones as extract_normalize_phones


GFORM_XLSX = Path(r"C:\RenStudio\case\washinmachine\顧客名單\淨新Google預約表單.xlsx")
ROWS_CSV = Path(__file__).parent / "out" / "rows.csv"
OUT_CSV = Path(__file__).parent / "out" / "gform_rows.csv"

# 與 extract.py 同步的 header
HEADERS = [
    "source_file", "sheet", "row_no", "format",
    "date", "name", "phones", "addresses",
    "brand_raw", "amount_raw", "item_raw", "payment", "source", "machine_id",
    "legacy_code", "staff_name", "actual_amount",
]


def normalize_phone(v) -> str:
    """單一電話正規化"""
    if v is None or v == "":
        return ""
    digits = re.sub(r"\D", "", str(v))
    if not digits:
        return ""
    # 9 開頭 9 碼 → 補 0
    if len(digits) == 9 and digits.startswith("9"):
        digits = "0" + digits
    # 太短跳過
    if len(digits) < 7:
        return ""
    return digits


def clean_brand(s) -> str:
    """gform 廠牌：『LG 樂金』『Panasonic 國際牌』『SAMPO 聲寶』→ 取第一段（英文）。
    若整段都中文（如『國際牌』）→ 整段保留。"""
    if not s:
        return ""
    s = str(s).strip()
    # 取第一段（空白前）
    first = re.split(r"\s+", s, maxsplit=1)[0]
    # 如果第一段是純英文/數字/連字號 → 用它；不然保留原文
    if re.fullmatch(r"[A-Za-z0-9\-]+", first):
        return first
    return s


def parse_qty(v) -> int:
    """『1台』『2台』『1』『2』→ 整數"""
    if v is None or v == "":
        return 1
    m = re.search(r"(\d+)", str(v))
    return int(m.group(1)) if m else 1


def main():
    print(f"讀 {GFORM_XLSX.name}")
    wb = load_workbook(str(GFORM_XLSX))
    sh = wb.worksheets[0]

    # 先載入 rows.csv 的 (phone, year-month) 集合，用來去重
    seen_keys = set()
    if ROWS_CSV.exists():
        with open(ROWS_CSV, encoding="utf-8-sig") as f:
            for r in csv.DictReader(f):
                phones = (r["phones"] or "").split("|")
                if not phones[0]:
                    continue
                dt = r["date"] or ""
                if len(dt) < 7:
                    continue
                seen_keys.add((phones[0], dt[:7]))
        print(f"  cleaned 已有 {len(seen_keys)} 個 (電話, 月份) 組合")
    else:
        print(f"  ⚠ {ROWS_CSV} 不存在，不去重")

    n_total = 0
    n_skipped_dup = 0
    n_skipped_blank = 0
    n_written = 0

    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=HEADERS)
        w.writeheader()

        for r in range(2, sh.max_row + 1):
            n_total += 1
            ts = sh.cell(r, 1).value
            name = sh.cell(r, 3).value
            phone_mobile = sh.cell(r, 4).value
            phone_home = sh.cell(r, 5).value
            address = sh.cell(r, 6).value
            brand_vertical = sh.cell(r, 7).value
            brand_drum = sh.cell(r, 8).value
            qty_wash = sh.cell(r, 9).value
            brand_ac = sh.cell(r, 10).value
            qty_ac = sh.cell(r, 11).value
            bed_sofa = sh.cell(r, 12).value
            note = sh.cell(r, 13).value
            source = sh.cell(r, 14).value

            # 空白列跳過
            if not any([name, phone_mobile, phone_home, address,
                       brand_vertical, brand_drum, brand_ac, bed_sofa]):
                n_skipped_blank += 1
                continue

            # 日期
            if not isinstance(ts, datetime):
                n_skipped_blank += 1
                continue
            date_str = ts.strftime("%Y-%m-%d")

            # 電話組合 — 用 extract.normalize_phones 處理「兩支電話沒分隔」case
            phones = []
            for cell in (phone_mobile, phone_home):
                for p in extract_normalize_phones(cell):
                    if p not in phones:
                        phones.append(p)
            if not phones:
                # 沒電話 → 不去重，但仍寫進 csv（之後 build.py 會分到 manual_review）
                pass
            primary_phone = phones[0] if phones else ""

            # 去重：跟 cleaned 同 (主電話, YYYY-MM) 就跳過
            if primary_phone:
                key = (primary_phone, date_str[:7])
                if key in seen_keys:
                    n_skipped_dup += 1
                    continue

            # 多機型 → 拆成多筆 row（每筆 1 個 machine_type）
            # 用 「+」 串接 brand_raw 讓 build.py 拆 items
            brand_parts = []
            if brand_vertical:
                b = clean_brand(brand_vertical)
                q = parse_qty(qty_wash) if qty_wash and not brand_drum else 1
                # 數量 N 用 *N 表示
                brand_parts.append(b + (f"*{q}" if q > 1 else ""))
            if brand_drum:
                b = clean_brand(brand_drum)
                q = parse_qty(qty_wash) if qty_wash and not brand_vertical else 1
                brand_parts.append(b + "(滾筒)" + (f"*{q}" if q > 1 else ""))
            if brand_ac:
                b = clean_brand(brand_ac)
                q = parse_qty(qty_ac)
                brand_parts.append(b + "(冷氣)" + (f"*{q}" if q > 1 else ""))
            if bed_sofa:
                brand_parts.append(str(bed_sofa).strip() + "(床沙)")

            brand_raw = "+".join(brand_parts) if brand_parts else ""

            w.writerow({
                "source_file": "Google預約表單.xlsx",
                "sheet": "表單回覆 1",
                "row_no": r,
                "format": "G",  # 新格式 G = Google form
                "date": date_str,
                "name": (str(name).strip() if name else ""),
                "phones": "|".join(phones),
                "addresses": (str(address).strip() if address else ""),
                "brand_raw": brand_raw,
                "amount_raw": "",  # gform 沒填金額
                "item_raw": "",
                "payment": "",
                "source": (str(source).strip() if source else ""),
                "machine_id": "",
                "legacy_code": "",
                "staff_name": "",
                "actual_amount": "",
            })
            n_written += 1

    print(f"\n總列數:      {n_total}")
    print(f"空白跳過:    {n_skipped_blank}")
    print(f"重複跳過:    {n_skipped_dup}")
    print(f"寫出:        {n_written}")
    print(f"\n→ {OUT_CSV}")

    # ---- 重寫 rows.csv（保留 cleaned 部分 + 新 gform 部分）----
    # 為避免重跑 append 兩次，先把 rows.csv 裡的 gform 部分移除，再 append。
    if ROWS_CSV.exists() and n_written > 0:
        with open(ROWS_CSV, encoding="utf-8-sig") as f:
            old_rows = [r for r in csv.DictReader(f)
                        if r["source_file"] != "Google預約表單.xlsx"]
        # 加上新的 gform rows
        with open(OUT_CSV, encoding="utf-8-sig") as src:
            new_rows = list(csv.DictReader(src))
        merged = old_rows + new_rows
        with open(ROWS_CSV, "w", encoding="utf-8-sig", newline="") as dst:
            writer = csv.DictWriter(dst, fieldnames=HEADERS)
            writer.writeheader()
            for r in merged:
                writer.writerow(r)
        print(f"\n已合併進 {ROWS_CSV.name}: {len(old_rows)} (cleaned) + {len(new_rows)} (gform) = {len(merged)} 筆")


if __name__ == "__main__":
    main()
