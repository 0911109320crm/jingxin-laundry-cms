"""
產生「客戶總覽」單一檔案 customers_overview.xlsx

每個客戶一列，所有資訊 join 在同一行。

分類成多張工作頁（每個客戶只進一張）：
  Sheet 1 ✅ 可信客戶          ← 第一次匯入優先選這批
  Sheet 2 ⚠ 電話可疑           ← 短於 9 碼 / 長於 11 碼 / 假電話
  Sheet 3 ⚠ 地址可疑           ← 縣市未分類 / 無『號樓巷』
  Sheet 4 ⚠ 無訂單紀錄         ← 只有資料沒服務紀錄
  Sheet 0 摘要                  ← 各 sheet 統計
"""
from __future__ import annotations
import csv
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "out"
REPORT = OUT / "customers_overview.xlsx"


def load(name):
    with open(OUT / name, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


print("讀 csv...")
customers = load("customers.csv")
phones = load("customer_phones.csv")
addresses = load("customer_addresses.csv")
machines = load("machines.csv")
orders = load("orders.csv")

print(f"  {len(customers)} 客戶 / {len(phones)} 電話 / {len(addresses)} 地址 / "
      f"{len(machines)} 機器 / {len(orders)} 訂單\n")

# 索引
phones_by_cust = defaultdict(list)
for p in phones:
    phones_by_cust[p["customer_id"]].append(p)

addrs_by_cust = defaultdict(list)
for a in addresses:
    addrs_by_cust[a["customer_id"]].append(a)

machines_by_cust = defaultdict(list)
for m in machines:
    machines_by_cust[m["customer_id"]].append(m)

orders_by_cust = defaultdict(list)
for o in orders:
    orders_by_cust[o["customer_id"]].append(o)

MTYPE_LABEL = {
    "washing_machine": "洗",
    "air_conditioner": "冷",
    "mattress": "床",
    "sofa": "沙",
    "other": "其",
}

import re

# ── 分類規則 ──────────────────────────────────────────────────────────────
def classify_customer(c, ps, a_list, n_orders):
    """傳回 ('ok'|'phone'|'address'|'no_order', 原因)"""
    # 1. 電話可疑：所有電話都 < 9 碼 或 > 11 碼 或 純假電話
    bad_phone_reasons = []
    valid_phone_count = 0
    for p in ps:
        d = re.sub(r"\D", "", p["phone"])
        if len(d) < 9:
            bad_phone_reasons.append(f"{p['phone']} 只 {len(d)} 碼")
            continue
        if len(d) > 11:
            bad_phone_reasons.append(f"{p['phone']} 有 {len(d)} 碼")
            continue
        # 純假電話：00000000、單調數字、12345678
        if len(set(d)) <= 2:
            bad_phone_reasons.append(f"{p['phone']} 數字單調")
            continue
        if d == "1234567890":
            bad_phone_reasons.append("測試假電話")
            continue
        valid_phone_count += 1
    if valid_phone_count == 0:
        return ("phone", "；".join(bad_phone_reasons) or "無有效電話")

    # 2. 地址可疑：所有地址都「未分類」或全無『號/樓/巷/弄/段/村/里』
    bad_addr_reasons = []
    valid_addr_count = 0
    for a in a_list:
        addr = a["address"]
        if a["county"] == "未分類" and a["district"] == "未分類":
            bad_addr_reasons.append(f"{addr} 縣市鄉鎮全未解析")
            continue
        if not re.search(r"號|樓|室|巷|弄|段|村|里", addr):
            bad_addr_reasons.append(f"{addr} 無號樓巷弄")
            continue
        valid_addr_count += 1
    if a_list and valid_addr_count == 0:
        return ("address", "；".join(bad_addr_reasons[:3]) or "地址無效")

    # 3. 無訂單
    if n_orders == 0:
        return ("no_order", "只有資料、無服務紀錄")

    return ("ok", "")


# ── 建構每個客戶的 row ─────────────────────────────────────────────────────
print("分類客戶...")
buckets = {
    "ok": [],
    "phone": [],
    "address": [],
    "no_order": [],
}

for c in sorted(customers, key=lambda x: x["code"]):
    cid = c["id"]
    ps = sorted(phones_by_cust[cid],
                key=lambda x: (0 if x["is_primary"] == "true" else 1, int(x.get("sort_order") or 0)))
    primary_phone = c["phone"]
    extra_phones = [p["phone"] + (f"({p['label']})" if p["label"] else "")
                    for p in ps if p["is_primary"] != "true"]

    a_list = addrs_by_cust[cid]
    default_addr = next((a for a in a_list if a["is_default"] == "true"), a_list[0] if a_list else None)
    default_str = (
        f"{default_addr['county']}{default_addr['district']} {default_addr['address']}"
        if default_addr else ""
    )
    other_addrs = [
        f"{a['county']}{a['district']} {a['address']}"
        for a in a_list if a is not default_addr
    ]

    m_list = machines_by_cust[cid]
    type_count = defaultdict(int)
    for m in m_list:
        type_count[m["type"]] += 1
    type_summary = " ".join(f"{MTYPE_LABEL.get(t, t)}×{n}" for t, n in type_count.items())
    machine_details = "; ".join(
        f"{MTYPE_LABEL.get(m['type'], m['type'])} {m['brand'] or '(無)'}"
        for m in m_list
    )

    o_list = sorted(orders_by_cust[cid], key=lambda x: x["service_at"])
    n_orders = len(o_list)
    first_date = o_list[0]["service_at"][:10] if o_list else ""
    last_date = o_list[-1]["service_at"][:10] if o_list else ""
    total_spent = sum(float(o["total"]) for o in o_list if o["total"])
    last_legacy = next((o["legacy_code"] for o in reversed(o_list) if o.get("legacy_code")), "")

    category, reason = classify_customer(c, ps, a_list, n_orders)

    row = [
        c["code"], c["name"], primary_phone,
        ", ".join(extra_phones),
        default_str,
        "; ".join(other_addrs),
        type_summary,
        machine_details[:150],
        n_orders,
        first_date, last_date,
        f"{total_spent:,.0f}" if total_spent else "",
        last_legacy,
        c.get("note", ""),
        c.get("joined_at", ""),
        reason,  # 暫緩原因（ok 為空）
    ]
    buckets[category].append(row)


# ── 寫入多 sheet ──────────────────────────────────────────────────────────
print("產出 xlsx (多 sheet)...")
wb = Workbook()
wb.remove(wb.active)

SHEET_DEFS = [
    ("摘要", None),
    ("✅ 可信客戶 (首批匯入)", "ok"),
    ("⚠ 電話可疑", "phone"),
    ("⚠ 地址可疑", "address"),
    ("⚠ 無訂單紀錄", "no_order"),
]

base_headers = [
    "編號", "姓名", "主電話", "副電話", "預設地址", "其他地址",
    "機器(數量)", "機器詳細", "訂單數", "首次服務", "末次服務",
    "累計金額", "舊系統清洗編號(最近)", "備註", "加入日期",
    "暫緩原因",
]
widths = [12, 12, 14, 28, 32, 25, 14, 30, 8, 12, 12, 11, 18, 18, 12, 30]

bold = Font(bold=True)
fill = PatternFill("solid", fgColor="EEEEEE")

# 摘要 sheet
sum_sh = wb.create_sheet("摘要")
sum_sh.append(["分類", "客戶數", "說明"])
sum_sh.append(["✅ 可信客戶 (首批匯入)", len(buckets["ok"]),
               "電話/地址/訂單都齊全，建議第一批就匯入"])
sum_sh.append(["⚠ 電話可疑", len(buckets["phone"]),
               "電話 < 9 碼、> 11 碼、單調數字（00000000、1234567890）"])
sum_sh.append(["⚠ 地址可疑", len(buckets["address"]),
               "地址縣市鄉鎮無法解析，或無『號/樓/巷/弄/段/村/里』"])
sum_sh.append(["⚠ 無訂單紀錄", len(buckets["no_order"]),
               "只有客戶資料（gform 預約）但沒實際服務紀錄"])
sum_sh.append([])
sum_sh.append(["總計", len(customers), ""])
sum_sh.append([])
sum_sh.append(["匯入順序建議:"])
sum_sh.append(["  1. 先用 import-legacy.mjs --filter=ok 只匯入「✅ 可信客戶」"])
sum_sh.append(["  2. 老闆娘上線試用一週後，再人工核對「⚠」三個 sheet"])
sum_sh.append(["  3. 核對完的逐批用 --filter=phone/address/no_order 匯入"])

for c in range(1, 4):
    sum_sh.cell(1, c).font = bold
    sum_sh.cell(1, c).fill = fill
sum_sh.column_dimensions["A"].width = 30
sum_sh.column_dimensions["B"].width = 12
sum_sh.column_dimensions["C"].width = 60

# 各分類 sheet
for sheet_name, bucket_key in SHEET_DEFS[1:]:
    sh = wb.create_sheet(sheet_name)
    sh.append(base_headers)
    for row in buckets[bucket_key]:
        sh.append(row)
    # 樣式
    for c in range(1, len(base_headers) + 1):
        sh.cell(1, c).font = bold
        sh.cell(1, c).fill = fill
        sh.cell(1, c).alignment = Alignment(horizontal="center", vertical="center")
    sh.freeze_panes = "C2"
    for i, w in enumerate(widths, start=1):
        sh.column_dimensions[get_column_letter(i)].width = w
    sh.auto_filter.ref = sh.dimensions

wb.save(str(REPORT))
print(f"\n→ {REPORT}")
print(f"\n各 sheet 客戶數:")
for k in ["ok", "phone", "address", "no_order"]:
    print(f"  {k}: {len(buckets[k])}")
